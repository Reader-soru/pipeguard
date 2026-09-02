import openpyxl
import os

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Team Blind Evaluation Dataset Subsea Pipeline.xlsx")

def load_scenario_telemetry(sheet_name: str):
    """
    Loads telemetry from a specific sheet in the Excel dataset.
    Returns a list of dicts: [{'timestamp': str, 'rel_time_ms': float, 'rel_time_s': float, 'inlet_p': float, 'outlet_p': float}]
    """
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in Excel dataset. Available: {wb.sheetnames}")
    
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    data = []
    
    # Row 0 is header: Timestamp, Relative Time (ms), Inlet Pressure (Bar), Outlet Pressure (Bar)
    for r in rows[1:]:
        if len(r) >= 4 and r[1] is not None:
            try:
                ts = str(r[0]) if r[0] is not None else ""
                rel_ms = float(r[1])
                rel_s = rel_ms / 1000.0
                inlet_p = float(r[2])
                outlet_p = float(r[3])
                data.append({
                    "timestamp": ts,
                    "rel_time_ms": rel_ms,
                    "rel_time_s": round(rel_s, 2),
                    "inlet_p": round(inlet_p, 4),
                    "outlet_p": round(outlet_p, 4)
                })
            except (ValueError, TypeError):
                continue
    return data
import csv
import io

def parse_csv_telemetry(csv_content: str):
    """
    Parses telemetry from a CSV string.
    Expected columns: Timestamp, Relative Time (ms), Inlet Pressure (Bar), Outlet Pressure (Bar)
    """
    reader = csv.reader(io.StringIO(csv_content))
    rows = list(reader)
    data = []
    for r in rows[1:]:
        if len(r) >= 4 and r[1]:
            try:
                ts = str(r[0])
                rel_ms = float(r[1])
                rel_s = rel_ms / 1000.0
                inlet_p = float(r[2])
                outlet_p = float(r[3])
                data.append({
                    "timestamp": ts,
                    "rel_time_ms": rel_ms,
                    "rel_time_s": round(rel_s, 2),
                    "inlet_p": round(inlet_p, 4),
                    "outlet_p": round(outlet_p, 4)
                })
            except (ValueError, TypeError):
                continue
    return data
def get_available_scenarios():
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    scenarios = [s for s in wb.sheetnames if s.startswith("BLIND_")]
    wb.close()
    return scenarios
